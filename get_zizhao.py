#coding:utf-8

#思路：
#首先获取到有哪些大学，并获取到这些大学的链接并注意编号 -- okay
#然后遍历每个大学，得到每个大学的名单列表 --okay
#最后把名单列表储存到Excel，并且对不同大学分表进行表达。

#下一步：
#1.多线程同时进行，太tm慢了
#2.增加对比不同功能，方便更新名单
#3.边爬边存，防止崩溃和内存占用过大

#imports:
import requests
from bs4 import BeautifulSoup
import re
import time, threading
import openpyxl
#settings:
threads = 5

#functions:
def get_colleges():
    #get the website
    colleges_website = requests.get("https://gaokao.chsi.com.cn/zzbm/mdgs/orgs.action")
    colleges_website.encoding = 'utf-8'
    colleges_website_content = BeautifulSoup(colleges_website.text, "html.parser").find_all("ul")[2].find_all("li")
    colleges_list = []
    college_id = 0
    for college in colleges_website_content:
        college_id += 1
        try:
            college["title"]
        except KeyError:
            college_type = 1
            college_oid = re.search(r"\d\d\d\d\d\d\d\d\d",re.search(r"oid=\d\d\d\d\d\d\d\d\d",college.a["href"]).group()).group()
            college_name = college.a.string.strip()
            colleges_list.append({"id" : college_id ,"type" : college_type ,"name" : college_name ,"oid" : college_oid})
        else:
            college_type = 0
            college_name = college.string.strip()
            colleges_list.append({"id" : college_id ,"type" : college_type ,"name" : college_name})
    colleges_num = college_id #the final college id.
    return [colleges_list,colleges_num]

def get_students(college_id,colleges):
    error = []
    students_information= []
    this_college = colleges[college_id]
    if this_college["type"] == 0 :
        students_information.append([{"type":0,"college_id" : college_id}])
    else:
        this_students_information = []
        this_oid = this_college["oid"]
        this_website = requests.get("https://gaokao.chsi.com.cn/zzbm/mdgs/detail.action", params = {"oid" : this_oid})
        this_pages_total = int(BeautifulSoup(this_website.text, "html.parser").find_all("span")[1].string)
        for this_page in range(this_pages_total):
            print("Scaaning the page" +str(this_page +1) + "/" + str(this_pages_total))
            try :
                this_website = requests.get("https://gaokao.chsi.com.cn/zzbm/mdgs/detail.action",params = {"oid": this_oid , "start" : (this_page)*30})
            except:
                error.append({"oid": this_oid , "start" : (this_page)*30})
            else:
                this_students_lists_html = BeautifulSoup(this_website.text, "html.parser").find_all("tr")
                this_students_id = 0
                for this_student in this_students_lists_html:
                    if this_students_id == 0:
                        pass
                    else:
                        this_student_information = {"type":1,
                                                    "college_id" : college_id,
                                                    "name":this_student.find_all("td")[0].string,
                                                        "sexy":this_student.find_all("td")[1].string,
                                                    "school":this_student.find_all("td")[2].string,
                                                    "province":this_student.find_all("td")[3].string}
                        this_students_information.append(this_student_information)
                    this_students_id += 1
        students_information.append(this_students_information)
    return students_information
            

def save_to_excel(students,colleges_name):
    excel = openpyxl.Workbook()
    sheet = excel.create_sheet("2018年自主招生报名审核通过名单")
    student_number = 1
    for i in range(6):
        sheet.cell(row = 1 , column = i+2).value = ["大学编号","大学名称","生源学校","学生姓名","性别","所在地"][i]
        #sheet.write(0,(i+1),["大学编号","大学名称","生源学校","学生姓名","性别","所在地"][i])
    for this_colleges in list(students):
        for this_student in list(this_colleges):
            print(this_student)
            if this_student["type"] == 0:
                sheet.cell(student_number+1, 1).value = str(student_number)
                sheet.cell(student_number+1, 2).value = '：该大学暂无名单'
                #sheet.write(student_number,1,"该大学暂无名单")
                student_number +=1
            else:
                for x in range(7):
                    #input("("+str(student_number)+","+str(x)+")")
                    sheet.cell(student_number+1,x+1).value = [str(student_number),str(this_student["college_id"]),str(colleges_name[(this_student["college_id"]-1)]["name"]),this_student["school"],this_student["name"],this_student["sexy"],this_student["province"]][x]
                student_number +=1
                print(student_number)
    excel.save("students.xlsx")

def save_to_file(students):
    x = open("./brain_data",'w',encoding='utf-8')
    x.write(str(students))
    x.close()
    file_student = open("./students.json","w",encoding='utf-8')
    for n in students:
        file_student.write(n)

def main():
    print("Now getting the colleges.")
    colleges = get_colleges()[0]
    print("Getting colleges finished.")
    students_total = []
    print("Now getting the studengts in colleges")
    colleges_num = get_colleges()[1]
    for college_id in range(colleges_num):
        print("The college id is " + str(college_id+1) + ".and it's name is " + colleges[college_id]["name"] +".Please wait for a moment."+"College number" + str(college_id+1) + "/" + str(colleges_num))
        this_students = get_students(college_id,colleges)
        for i in this_students:
            students_total.append(i)
    print("Getting students finished.")
    print("Start to saving it to Excel.")
    save_to_file(students_total)
    save_to_excel(students_total,colleges)
    print("Now start to save into file.")
    
    print("All have done.")
    

if __name__ == "__main__":
    main()
    