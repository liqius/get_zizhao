#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

from get_zizhao import get_colleges


def save_to_excel(students,colleges_name):
    excel = openpyxl.Workbook()
    sheet = excel.create_sheet("2018年自主招生报名审核通过名单")
    student_number = 1
    for i in range(6):
        sheet.cell(row = 1 , column = i+2).value = ["大学编号","大学名称","生源学校","学生姓名","性别","所在地"][i]
        #sheet.write(0,(i+1),["大学编号","大学名称","生源学校","学生姓名","性别","所在地"][i])
    for this_colleges in list(students):
        for this_student in list(this_colleges):
            print(this_student)
            if this_student["type"] == 0:
                sheet.cell(student_number+1, 1).value = str(student_number)
                sheet.cell(student_number+1, 2).value = '：该大学暂无名单'
                #sheet.write(student_number,1,"该大学暂无名单")
                student_number +=1
            else:
                for x in range(7):
                    #input("("+str(student_number)+","+str(x)+")")
                    sheet.cell(student_number+1,x+1).value = [str(student_number),str(this_student["college_id"]),str(colleges_name[(this_student["college_id"]-1)]["name"]),this_student["school"],this_student["name"],this_student["sexy"],this_student["province"]][x]
                student_number +=1
                print(student_number)
    excel.save("students.xlsx")

i = open("./brain_data","r",encoding='utf-8')
n = eval(i.read())
#print(n)
l = []
#print(type(n))
#save_to_excel(n,get_colleges()[0])
for m in n:
    for b in m:
        if b['college_id'] in l and b['type'] == 1:
            #print(b['type'])
            #input(str(b))
            pass
        elif b['type'] == 1:
            l.append(b["college_id"])
print(l)
print(len(l))